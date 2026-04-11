"""
Excel Parser — parses individual project Gantt Excel files.
Same format as 5001. Returns (project_info, tasks, dependencies).
"""

import openpyxl
from datetime import datetime, timedelta
import re
import config


class ExcelParser:
    def parse_excel_file(self, file_path):
        print(f'[Parser] Parsing: {file_path}')
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active

            project_info = {
                'name':           str(ws['C3'].value or '').strip(),
                'manager':        str(ws['C5'].value or '').strip(),
                'excel_filename': str(file_path),
            }
            if not project_info['name']:
                import os
                project_info['name'] = os.path.splitext(os.path.basename(str(file_path)))[0]

            print(f'  Project: {project_info["name"]}')
            print(f'  Manager: {project_info["manager"]}')

            tasks = []
            row_order = 1
            for row_num in range(config.EXCEL_DATA_START_ROW, ws.max_row + 1):
                row = ws[row_num]
                ref_id      = row[0].value
                description = row[1].value
                source      = row[2].value
                assigned_to = row[3].value
                date_entered = row[4].value
                status      = row[5].value
                due_date    = row[6].value
                date_closed = row[7].value
                result      = row[8].value

                if not description:
                    continue

                desc_str = str(description).strip()
                is_gate = bool(re.match(r'^Gate\s*\d+\s*$', desc_str, re.IGNORECASE))

                start_str = self._parse_date(date_entered)
                end_str   = self._parse_date(due_date)

                if not end_str and start_str:
                    try:
                        end_str = (datetime.strptime(start_str, '%Y-%m-%d') + timedelta(days=30)).strftime('%Y-%m-%d')
                    except Exception:
                        end_str = ''

                tasks.append({
                    'reference_id': str(ref_id or '').strip(),
                    'name':         desc_str,
                    'phase':        str(source or '').strip(),
                    'owner':        str(assigned_to or '').strip().split(',')[0].strip(),
                    'start_date':   start_str,
                    'end_date':     end_str,
                    'status':       str(status or 'Planned').strip(),
                    'date_closed':  self._parse_date(date_closed) if date_closed else '',
                    'result':       str(result or '').strip(),
                    'milestone':    1 if is_gate else 0,
                    'row_order':    row_order,
                })
                row_order += 1

            print(f'  Tasks parsed: {len(tasks)}')
            wb.close()
            return project_info, tasks

        except Exception as e:
            print(f'  [ERROR] {e}')
            raise

    @staticmethod
    def _parse_date(val):
        if isinstance(val, datetime):
            return val.strftime('%Y-%m-%d')
        if isinstance(val, str) and val.strip():
            for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d'):
                try:
                    return datetime.strptime(val.strip(), fmt).strftime('%Y-%m-%d')
                except ValueError:
                    pass
        return ''
