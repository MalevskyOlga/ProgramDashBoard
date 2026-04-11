"""
Priority Parser — reads Flame & Gas Project Priority List Excel.
Targets the latest sheet (highest date tab, e.g. 'Mar26').
Returns a list of project dicts ready for upsert into portfolio_projects.
"""

import openpyxl
from pathlib import Path


# Map Excel column letters (0-based index) for the Big Projects section
_COL = {
    'priority':         0,   # A
    'name':             1,   # B
    'leader':           2,   # C
    'process_type':     3,   # D
    'next_gate':        4,   # E
    'planned_launch':   5,   # F
    'objective':        6,   # G
    'business_segment': 7,   # H
}

# Resource discipline columns start at index 9 (J) in the Mar26 sheet
# coverage block:  cols 9–24  (R&D, ENG, Ops, Supply Chain, Other)
# demand block:    cols 25–40
_DISCIPLINE_NAMES = [
    'Proj. Mgmt', 'Optics', 'EE', 'ME', 'SW DEV', 'R&D QA',
    'Sderot Cert.', 'RSK Cert.', 'ATE', 'NPD & Main.', 'FCT',
    'RSK Ops', 'Sderot', 'RSK', 'Product Mgmt',
]

_COVERAGE_START = 9    # column index (0-based) of first coverage column
_DEMAND_START   = 25   # column index of first demand column (Mar26 layout)

# Valid coverage / demand tokens
_COVERAGE_VALID = {'Fully', 'Partially', 'No', 'N/A'}
_DEMAND_VALID   = {'Large', 'Medium', 'Small', 'N/A'}

# Month abbreviations for sheet sorting
_MONTHS = {'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4, 'May': 5, 'Jun': 6,
           'Jul': 7, 'July': 7, 'Aug': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12}


def _sheet_sort_key(name):
    """Turn 'Mar26' → (2026, 3) for sorting newest-first."""
    for mon, num in _MONTHS.items():
        if name.startswith(mon):
            try:
                yr = int(name[len(mon):])
                return (2000 + yr if yr < 100 else yr, num)
            except ValueError:
                pass
    return (0, 0)


def parse_priority_list(file_path, sheet_name=None):
    """
    Parse the priority list Excel file.

    Args:
        file_path: path to the .xlsx file
        sheet_name: specific sheet to use; if None, picks the latest dated sheet

    Returns:
        dict with keys:
          'sheet'         : sheet name used
          'big_projects'  : list of project dicts (ACTIVE Big Projects)
          'small_projects': list of small project dicts
          'on_hold'       : list of on-hold project dicts
          'proposed'      : list of proposed project dicts
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Pick sheet
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        chosen = sheet_name
    else:
        dated = [s for s in wb.sheetnames if _sheet_sort_key(s) != (0, 0)]
        if dated:
            chosen = max(dated, key=_sheet_sort_key)
        else:
            chosen = wb.sheetnames[0]
        ws = wb[chosen]

    rows = list(ws.iter_rows(values_only=True))

    big_projects   = []
    small_projects = []
    on_hold        = []
    proposed       = []

    current_section = None
    small_header_seen = False

    for row in rows:
        if not any(c is not None for c in row):
            continue

        # Detect section headers
        cell_b = str(row[1] or '').strip()
        if 'ACTIVE Big Projects' in cell_b:
            current_section = 'big'
            continue
        if 'ACTIVE Small Projects' in cell_b:
            current_section = 'small'
            small_header_seen = False
            continue
        if 'PLANNED' in cell_b and 'ON HOLD' in cell_b:
            current_section = 'on_hold'
            continue
        if 'PROPOSED' in cell_b:
            current_section = 'proposed'
            continue

        # Skip header rows
        cell_a = str(row[0] or '').strip()
        if cell_a.lower() in ('priority', 'none') or cell_b.lower() in ('project name', ''):
            continue

        if current_section == 'big':
            p = _parse_big_row(row)
            if p:
                big_projects.append(p)

        elif current_section == 'small':
            # Small projects header row has 'Project Name' in col B
            if cell_b.lower() == 'project name':
                small_header_seen = True
                continue
            if cell_b and cell_b.lower() not in ('project name',):
                p = _parse_small_row(row)
                if p:
                    small_projects.append(p)

        elif current_section == 'on_hold':
            if cell_b and cell_b.lower() not in ('project name', 'remarks'):
                on_hold.append({'name': cell_b,
                                'business_segment': str(row[2] or '').strip(),
                                'remarks': str(row[3] or '').strip()})

        elif current_section == 'proposed':
            if cell_b and cell_b.lower() not in ('project name', 'remarks'):
                proposed.append({'name': cell_b,
                                 'business_segment': str(row[2] or '').strip(),
                                 'remarks': str(row[3] or '').strip()})

    wb.close()
    return {
        'sheet':          chosen,
        'big_projects':   big_projects,
        'small_projects': small_projects,
        'on_hold':        on_hold,
        'proposed':       proposed,
    }


def _parse_big_row(row):
    """Parse a single big-project row."""
    priority_raw = row[_COL['priority']]
    name_raw     = row[_COL['name']]

    # Must have a numeric priority and a name
    try:
        priority = int(priority_raw)
    except (TypeError, ValueError):
        return None
    if not name_raw:
        return None

    name = str(name_raw).strip()
    if not name:
        return None

    project = {
        'priority':         priority,
        'name':             name,
        'leader':           str(row[_COL['leader']]         or '').strip(),
        'process_type':     str(row[_COL['process_type']]   or '').strip(),
        'next_gate':        str(row[_COL['next_gate']]      or '').strip(),
        'planned_launch':   str(row[_COL['planned_launch']] or '').strip(),
        'objective':        str(row[_COL['objective']]      or '').strip(),
        'business_segment': str(row[_COL['business_segment']] or '').strip(),
        'management_type':  'card',    # default; user upgrades to gantt in portal
        'resources':        {},        # discipline → {coverage, demand}
    }

    # Parse coverage columns
    for i, disc in enumerate(_DISCIPLINE_NAMES):
        col_idx = _COVERAGE_START + i
        cov_raw = str(row[col_idx] or '').strip() if col_idx < len(row) else ''
        coverage = cov_raw if cov_raw in _COVERAGE_VALID else 'N/A'

        dem_idx  = _DEMAND_START + i
        dem_raw  = str(row[dem_idx] or '').strip() if dem_idx < len(row) else ''
        demand   = dem_raw if dem_raw in _DEMAND_VALID else 'N/A'

        # Convert demand bucket → days (midpoint estimate)
        demand_days = _demand_to_days(demand)

        project['resources'][disc] = {
            'coverage':    coverage,
            'demand':      demand,     # original label
            'demand_days': demand_days,
        }

    return project


def _parse_small_row(row):
    """Parse a small project row (simpler schema)."""
    name = str(row[1] or '').strip()
    if not name:
        return None
    return {
        'name':             name,
        'leader':           str(row[2] or '').strip(),
        'objective':        str(row[3] or '').strip(),
        'business_segment': str(row[4] or '').strip(),
        'main_resource':    str(row[5] or '').strip(),
        'due_date':         str(row[6] or '').strip(),
        'management_type':  'card',
    }


def _demand_to_days(demand_label):
    """Convert demand bucket label to midpoint working days."""
    mapping = {
        'Large':  24.0,   # >16 days → use 24 as representative
        'Medium': 10.0,   # 5–16 days → use 10
        'Small':   2.5,   # <5 days  → use 2.5
        'N/A':     0.0,
    }
    return mapping.get(demand_label, 0.0)
