"""
Excel Exporter for Dashboard Generator Web Server
Exports dashboard data from database back to Excel format
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import os
import config


class ExcelExporter:
    def __init__(self):
        pass
    
    def export_project(self, project, tasks, output_folder):
        """
        Export project and tasks to Excel file
        
        Args:
            project: Project dict from database
            tasks: List of task dicts from database
            output_folder: Where to save the Excel file
            
        Returns:
            str: Path to generated Excel file
        """
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Action Items"
            
            # Set column widths
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 12
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 12
            ws.column_dimensions['H'].width = 12
            ws.column_dimensions['I'].width = 30
            
            # Header styles
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font_white = Font(bold=True, size=11, color="FFFFFF")
            
            # Write project header
            ws['A1'] = "Project Dashboard"
            ws['A1'].font = Font(bold=True, size=14)
            
            ws['B3'] = "Project Name:"
            ws['B3'].font = header_font
            ws['C3'] = project['name']
            
            ws['B5'] = "Project Manager:"
            ws['B5'].font = header_font
            ws['C5'] = project['manager']
            
            ws['B7'] = "Exported:"
            ws['B7'].font = header_font
            ws['C7'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Column headers (row 10)
            headers = [
                'Reference ID#',
                'Description of Action Item',
                'Source of Action Item',
                'Assigned to',
                'Date Entered',
                'Status',
                'Due Date',
                'Date Closed',
                'Final Result/Outcome'
            ]
            
            for col_num, header in enumerate(headers, start=1):
                cell = ws.cell(row=10, column=col_num)
                cell.value = header
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Write tasks (starting from row 11)
            for idx, task in enumerate(tasks, start=11):
                ws.cell(row=idx, column=1, value=task['reference_id'])
                ws.cell(row=idx, column=2, value=task['name'])
                ws.cell(row=idx, column=3, value=task['phase'])
                ws.cell(row=idx, column=4, value=task['owner'])
                ws.cell(row=idx, column=5, value=task['start_date'])
                ws.cell(row=idx, column=6, value=task['status'])
                ws.cell(row=idx, column=7, value=task['end_date'])
                ws.cell(row=idx, column=8, value=task['date_closed'])
                ws.cell(row=idx, column=9, value=task['result'])
            
            # Generate filename with timestamp (dd-mm-yy_hhmm format)
            timestamp = datetime.now().strftime('%d-%m-%y_%H%M')
            filename = f"Action Item List _{project['name']}_{timestamp}.xlsx"
            
            # Save file
            os.makedirs(output_folder, exist_ok=True)
            file_path = os.path.join(output_folder, filename)
            wb.save(file_path)
            
            print(f"✓ Exported to: {filename}")
            
            return file_path
            
        except Exception as e:
            print(f"✗ Error exporting to Excel: {e}")
            return None
