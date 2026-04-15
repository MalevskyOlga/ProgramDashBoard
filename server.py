"""
Dashboard Generator Web Server
Flask-based web server for managing project dashboards
"""

import os
import sys
import tempfile
from datetime import datetime
from pathlib import Path

# Ensure stdout/stderr use UTF-8 so Unicode characters (✓ ✗ etc.) don't crash on Windows
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ('utf-8', 'utf8'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if sys.stderr.encoding and sys.stderr.encoding.lower() not in ('utf-8', 'utf8'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')
from flask import Flask, render_template, jsonify, request, send_file
from werkzeug.utils import secure_filename
from database_manager import DatabaseManager
from excel_parser import ExcelParser
import config
import sqlite3

app = Flask(__name__)
app.config['SECRET_KEY'] = 'dashboard-generator-secret-key'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Initialize database manager
db_manager = DatabaseManager(config.DATABASE_PATH)
db_manager.initialize_database()


def allowed_file(filename):
    """Check if uploaded file has .xlsx extension"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() == 'xlsx'


@app.after_request
def add_no_cache_headers(response):
    """Force fresh dashboard/API reads so detailed gantt edits propagate everywhere."""
    if request.method == 'GET':
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    # Log every request so errors are visible in server.stdout.log
    ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    status = response.status_code
    if status >= 400:
        # Log errors with response body so the cause is clear
        try:
            body = response.get_data(as_text=True)[:500]
        except Exception:
            body = ''
        print(f'[{ts}] {request.method} {request.path} -> {status} | {body}', flush=True)
    else:
        print(f'[{ts}] {request.method} {request.path} -> {status}', flush=True)
    return response


@app.route('/')
def index():
    """Home page - list all available projects"""
    projects = db_manager.get_all_projects()
    return render_template('index.html', projects=projects)


@app.route('/project/<project_name>')
def view_project(project_name):
    """View a specific project dashboard"""
    project = db_manager.get_project_by_name(project_name)
    if not project:
        return f"Project '{project_name}' not found", 404
    
    return render_template('dashboard.html', 
                         project_name=project_name,
                         project_manager=project['manager'],
                         PROJECT_NAME=project_name,
                         PROJECT_MANAGER=project['manager'],
                         GENERATED_DATE=datetime.now().strftime('%Y-%m-%d %H:%M'))


@app.route('/project/<project_name>/schematic')
def view_project_schematic(project_name):
    """View a standalone schematic schedule for a project"""
    project = db_manager.get_project_by_name(project_name)
    if not project:
        return f"Project '{project_name}' not found", 404

    return render_template(
        'project_schematic.html',
        project_name=project_name,
        project_manager=project['manager'],
        PROJECT_NAME=project_name,
        PROJECT_MANAGER=project['manager'],
        GENERATED_DATE=datetime.now().strftime('%Y-%m-%d %H:%M')
    )


@app.route('/api/projects')
def api_get_projects():
    """API endpoint to get all projects"""
    projects = db_manager.get_all_projects()
    return jsonify(projects)


@app.route('/api/overall-gate-timeline')
def api_get_overall_gate_timeline():
    """API endpoint to get unified gate timeline across all projects"""
    timeline = db_manager.get_overall_gate_timeline()
    return jsonify(timeline)


@app.route('/api/overall-resource-load')
def api_get_overall_resource_load():
    """API endpoint to get unified owner workload across all projects"""
    resource_load = db_manager.get_overall_resource_load()
    return jsonify(resource_load)


@app.route('/api/overall-critical-path-overview')
def api_get_overall_critical_path_overview():
    """API endpoint to get unified critical path overview across all projects"""
    critical_overview = db_manager.get_overall_critical_path_overview()
    return jsonify(critical_overview)


@app.route('/api/project/<project_name>')
def api_get_project(project_name):
    """API endpoint to get project details"""
    project = db_manager.get_project_by_name(project_name)
    if not project:
        return jsonify({'error': 'Project not found'}), 404
    
    return jsonify(project)


@app.route('/api/project/<project_name>/tasks')
def api_get_tasks(project_name):
    """API endpoint to get all tasks for a project"""
    tasks = db_manager.get_tasks_by_project(project_name)
    return jsonify(tasks)


@app.route('/api/task/<int:task_id>', methods=['GET'])
def api_get_task(task_id):
    """API endpoint to get a specific task"""
    task = db_manager.get_task_by_id(task_id)
    if not task:
        return jsonify({'error': 'Task not found'}), 404
    
    return jsonify(task)


@app.route('/api/task/<int:task_id>', methods=['PUT'])
def api_update_task(task_id):
    """API endpoint to update a task"""
    data = request.json
    
    # Validate required fields
    if not data:
        return jsonify({'error': 'No data provided'}), 400
    
    success = db_manager.update_task(task_id, data)
    
    if success:
        task = db_manager.get_task_by_id(task_id)
        return jsonify({'message': 'Task updated successfully', 'task': task})
    else:
        return jsonify({'error': 'Failed to update task'}), 500


@app.route('/api/project/<project_name>/task', methods=['POST'])
def api_create_task(project_name):
    """API endpoint to create a new task"""
    data = request.json
    
    # Validate required fields
    required_fields = ['name', 'phase', 'owner', 'start_date', 'end_date', 'status']
    for field in required_fields:
        if field not in data:
            return jsonify({'error': f'Missing required field: {field}'}), 400
    
    # Get project ID
    project = db_manager.get_project_by_name(project_name)
    if not project:
        return jsonify({'error': 'Project not found'}), 404
    
    data['project_id'] = project['id']
    
    task_id = db_manager.create_task(data)
    
    if task_id:
        task = db_manager.get_task_by_id(task_id)
        return jsonify({'message': 'Task created successfully', 'task': task}), 201
    else:
        return jsonify({'error': 'Failed to create task'}), 500


@app.route('/api/task/<int:task_id>', methods=['DELETE'])
def api_delete_task(task_id):
    """API endpoint to delete a task"""
    success = db_manager.delete_task(task_id)
    
    if success:
        return jsonify({'message': 'Task deleted successfully'})
    else:
        return jsonify({'error': 'Failed to delete task'}), 500


@app.route('/api/upload-excel', methods=['POST'])
def api_upload_excel():
    """API endpoint to upload and import Excel file"""
    temp_file_path = None
    
    try:
        # Check if file is present in request
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        
        # Check if file was actually selected
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        # Validate file extension
        if not allowed_file(file.filename):
            return jsonify({'error': 'Only .xlsx files are allowed'}), 400
        
        # Secure the filename
        filename = secure_filename(file.filename)
        
        # Create temporary file to store upload
        temp_fd, temp_file_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(temp_fd)
        
        # Save uploaded file to temporary location
        file.save(temp_file_path)
        
        # Initialize parser
        parser = ExcelParser()
        
        # Parse Excel file
        project_info, tasks = parser.parse_excel_file(temp_file_path)

        # Allow caller to override the project name (e.g. from priority list row)
        override_name = request.form.get('project_name', '').strip()
        if override_name:
            project_info['name'] = override_name

        # Create or update project
        project_id = db_manager.create_or_update_project(
            name=project_info['name'],
            manager=project_info['manager'],
            excel_filename=filename
        )
        
        # Delete existing tasks for this project
        deleted_count = db_manager.delete_tasks_by_project(project_id)
        if deleted_count > 0:
            print(f"  ✓ Removed {deleted_count} old tasks")
        
        # Import all tasks
        imported_count = 0
        for task in tasks:
            task['project_id'] = project_id
            task_id = db_manager.create_task(task)
            if task_id:
                imported_count += 1
        
        print(f"  ✓ Imported {imported_count} tasks for project '{project_info['name']}'")
        
        # Auto-detect dependencies by date matching (Finish-to-Start, delta <= 2 days)
        # If 0 <= task_B.start_date - task_A.end_date <= 2 days, they are dependent
        from datetime import datetime, timedelta
        db_manager.delete_dependencies_by_project(project_info['name'])
        dep_count = 0
        all_tasks_for_project = db_manager.get_tasks_by_project(project_info['name'])
        DEP_DELTA_DAYS = 2
        # Parse all end dates once for efficiency
        task_end_dates = {}
        for t in all_tasks_for_project:
            ed = t.get('end_date', '')
            if ed:
                try:
                    task_end_dates[t['id']] = datetime.strptime(ed, '%Y-%m-%d')
                except ValueError:
                    pass
        # For each task, find predecessors whose end_date is within DEP_DELTA_DAYS before this task's start_date
        for successor in all_tasks_for_project:
            sd = successor.get('start_date', '')
            if not sd:
                continue
            try:
                start_dt = datetime.strptime(sd, '%Y-%m-%d')
            except ValueError:
                continue
            for predecessor in all_tasks_for_project:
                if predecessor['id'] == successor['id']:
                    continue
                pred_end_dt = task_end_dates.get(predecessor['id'])
                if pred_end_dt is None:
                    continue
                delta = (start_dt - pred_end_dt).days
                if 0 <= delta <= DEP_DELTA_DAYS:
                    db_manager.create_dependency(project_info['name'], predecessor['id'], successor['id'])
                    dep_count += 1
        print(f"  ✓ Detected {dep_count} task dependencies (Finish-to-Start, delta <= {DEP_DELTA_DAYS} days)")
        
        return jsonify({
            'success': True,
            'message': f'Successfully imported {filename}',
            'project_name': project_info['name'],
            'tasks_imported': imported_count,
            'dependencies_detected': dep_count,
            'filename': filename
        }), 200
            
    except Exception as e:
        return jsonify({'error': f'Error processing file: {str(e)}'}), 500
        
    finally:
        # Clean up temporary file
        if temp_file_path and os.path.exists(temp_file_path):
            try:
                os.unlink(temp_file_path)
            except Exception as e:
                print(f"Warning: Failed to delete temporary file {temp_file_path}: {e}")


@app.route('/api/project/<project_name>/export')
def api_export_to_excel(project_name):
    """API endpoint to export project to Excel"""
    from excel_exporter import ExcelExporter
    
    project = db_manager.get_project_by_name(project_name)
    if not project:
        return jsonify({'error': 'Project not found'}), 404
    
    tasks = db_manager.get_tasks_by_project(project_name)
    
    exporter = ExcelExporter()
    excel_path = exporter.export_project(project, tasks, config.EXCEL_OUTPUT_FOLDER)
    
    if excel_path and os.path.exists(excel_path):
        return send_file(
            excel_path,
            as_attachment=True,
            download_name=os.path.basename(excel_path),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    else:
        return jsonify({'error': 'Failed to export to Excel'}), 500


@app.route('/api/project/<project_name>/export-ppt')
def api_export_schematic_to_ppt(project_name):
    """API endpoint to export project schematic to PowerPoint"""
    from ppt_exporter import PptExporter

    project = db_manager.get_project_by_name(project_name)
    if not project:
        return jsonify({'error': 'Project not found'}), 404

    tasks = db_manager.get_tasks_by_project(project_name)
    gate_sign_offs = db_manager.get_gate_sign_offs(project_name)

    exporter = PptExporter()
    ppt_path = exporter.export_schematic(project, tasks, gate_sign_offs, config.EXCEL_OUTPUT_FOLDER)

    if ppt_path and os.path.exists(ppt_path):
        return send_file(
            ppt_path,
            as_attachment=True,
            download_name=os.path.basename(ppt_path),
            mimetype='application/vnd.openxmlformats-officedocument.presentationml.presentation'
        )

    return jsonify({'error': 'Failed to export schematic to PowerPoint'}), 500


@app.route('/api/stats')
def api_get_stats():
    """API endpoint to get server statistics"""
    resource_load = db_manager.get_overall_resource_load()
    stats = {
        'total_projects': len(db_manager.get_all_projects()),
        'total_tasks': db_manager.get_total_task_count(),
        'total_owners': resource_load['summary']['owner_count'],
        'last_scan': 'N/A'
    }
    return jsonify(stats)


# Priority Projects Endpoints (DB-backed; Excel is read-only seed)

_PRIORITY_ALLOWED_FIELDS = {
    'priority', 'name', 'leader', 'process_type',
    'next_gate', 'launch_date', 'objective', 'segment', 'linked_dashboard'
}


def _ensure_priority_tables():
    """Return an open connection with priority tables created if absent."""
    conn = db_manager.get_connection()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS priority_projects (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            priority      INTEGER NOT NULL DEFAULT 0,
            name          TEXT NOT NULL DEFAULT '',
            leader        TEXT DEFAULT '',
            process_type  TEXT DEFAULT '',
            next_gate     TEXT DEFAULT '',
            launch_date   TEXT DEFAULT '',
            objective     TEXT DEFAULT '',
            segment       TEXT DEFAULT '',
            created_at    TEXT DEFAULT (datetime('now')),
            updated_at    TEXT DEFAULT (datetime('now'))
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS archived_projects (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            name           TEXT NOT NULL DEFAULT '',
            completed_date TEXT NOT NULL,
            priority       INTEGER DEFAULT 0,
            leader         TEXT DEFAULT '',
            process_type   TEXT DEFAULT '',
            next_gate      TEXT DEFAULT '',
            launch_date    TEXT DEFAULT '',
            objective      TEXT DEFAULT '',
            segment        TEXT DEFAULT ''
        )
    """)
    conn.commit()
    # Migration: add linked_dashboard column if it doesn't exist yet
    try:
        conn.execute("ALTER TABLE priority_projects ADD COLUMN linked_dashboard TEXT DEFAULT NULL")
        conn.commit()
    except Exception:
        pass  # column already exists
    return conn


def _seed_priority_from_excel(conn):
    """Seed priority_projects from Excel (called once when table is empty)."""
    try:
        import openpyxl
        excel_path = config.PRIORITY_LIST_PATH
        if not excel_path.exists():
            return
        wb = openpyxl.load_workbook(str(excel_path), data_only=True)
        ws = wb[config.PRIORITY_LIST_SHEET]
        for row_idx in range(1, ws.max_row + 1):
            pv = ws.cell(row=row_idx, column=1).value
            if not isinstance(pv, int) or pv < 1:
                continue
            conn.execute("""
                INSERT INTO priority_projects
                    (priority, name, leader, process_type, next_gate, launch_date, objective, segment)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                pv,
                str(ws.cell(row=row_idx, column=2).value or '').strip(),
                str(ws.cell(row=row_idx, column=3).value or '').strip(),
                str(ws.cell(row=row_idx, column=4).value or '').strip(),
                str(ws.cell(row=row_idx, column=5).value or '').strip(),
                str(ws.cell(row=row_idx, column=6).value or '').strip(),
                str(ws.cell(row=row_idx, column=7).value or '').strip(),
                str(ws.cell(row=row_idx, column=8).value or '').strip(),
            ))
        conn.commit()
    except Exception as e:
        print(f'Warning: could not seed priority list from Excel: {e}')


def _build_priority_response(conn):
    """Return sorted priority projects list with dashboard link hints."""
    # Only include projects that are NOT soft-deleted
    existing = {p['name'].strip().lower(): p['name'] for p in db_manager.get_all_projects()}
    rows = conn.execute(
        "SELECT * FROM priority_projects ORDER BY priority, id"
    ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        linked = d.get('linked_dashboard')
        if linked == '__none__':
            # User explicitly deleted the Gantt — suppress auto-match
            d['dashboard_name'] = None
        elif linked and linked.strip().lower() in existing:
            # Manual link — only if the project actually exists (not soft-deleted)
            d['dashboard_name'] = linked
        else:
            # Fall back to auto-match by row name (only when never explicitly linked/unlinked)
            d['dashboard_name'] = existing.get((d.get('name') or '').strip().lower())
        result.append(d)
    return result


@app.route('/api/priority-projects', methods=['GET'])
def api_get_priority_projects():
    """Return all active priority projects (seed from Excel on first call)."""
    try:
        conn = _ensure_priority_tables()
        if conn.execute("SELECT COUNT(*) FROM priority_projects").fetchone()[0] == 0:
            _seed_priority_from_excel(conn)
        result = _build_priority_response(conn)
        conn.close()
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/priority-projects', methods=['POST'])
def api_create_priority_project():
    """Append a blank project row at max priority + 1."""
    try:
        conn = _ensure_priority_tables()
        max_p = conn.execute(
            "SELECT COALESCE(MAX(priority), 0) FROM priority_projects"
        ).fetchone()[0]
        new_p = max_p + 1
        row_id = conn.execute("""
            INSERT INTO priority_projects
                (priority, name, leader, process_type, next_gate, launch_date, objective, segment)
            VALUES (?, '', '', '', '', '', '', '')
        """, (new_p,)).lastrowid
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'id': row_id, 'priority': new_p}), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/priority-projects/<int:proj_id>', methods=['PUT'])
def api_update_priority_project(proj_id):
    """Update one field of a priority project; cascade on priority conflicts."""
    try:
        data = request.json or {}
        field = data.get('field')
        value = data.get('value', '')

        if field not in _PRIORITY_ALLOWED_FIELDS:
            return jsonify({'error': f'Unknown field: {field}'}), 400

        conn = _ensure_priority_tables()

        if field == 'priority':
            try:
                new_p = int(value)
            except (ValueError, TypeError):
                conn.close()
                return jsonify({'error': 'Priority must be a number'}), 400
            if new_p < 1:
                conn.close()
                return jsonify({'error': 'Priority must be ≥ 1'}), 400

            row = conn.execute(
                "SELECT priority FROM priority_projects WHERE id = ?", (proj_id,)
            ).fetchone()
            if not row:
                conn.close()
                return jsonify({'error': 'Project not found'}), 404
            old_p = row[0]

            if new_p != old_p:
                conflict = conn.execute(
                    "SELECT id FROM priority_projects WHERE priority = ? AND id != ?",
                    (new_p, proj_id)
                ).fetchone()
                if conflict:
                    if new_p < old_p:  # moving up → shift others down
                        conn.execute("""
                            UPDATE priority_projects
                            SET priority = priority + 1
                            WHERE priority >= ? AND priority < ? AND id != ?
                        """, (new_p, old_p, proj_id))
                    else:              # moving down → shift others up
                        conn.execute("""
                            UPDATE priority_projects
                            SET priority = priority - 1
                            WHERE priority > ? AND priority <= ? AND id != ?
                        """, (old_p, new_p, proj_id))
            conn.execute(
                "UPDATE priority_projects SET priority = ?, updated_at = datetime('now') WHERE id = ?",
                (new_p, proj_id)
            )
        elif field == 'linked_dashboard':
            # Store '__none__' sentinel as-is; empty string → NULL; real name → real name
            stored = value if value else None
            conn.execute(
                "UPDATE priority_projects SET linked_dashboard = ?, updated_at = datetime('now') WHERE id = ?",
                (stored, proj_id)
            )
        else:
            conn.execute(
                f"UPDATE priority_projects SET {field} = ?, updated_at = datetime('now') WHERE id = ?",
                (value or '', proj_id)
            )

        conn.commit()
        result = _build_priority_response(conn)
        conn.close()
        return jsonify({'success': True, 'projects': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/priority-projects/<int:proj_id>', methods=['DELETE'])
def api_delete_priority_project(proj_id):
    """Permanently remove a priority project."""
    try:
        conn = _ensure_priority_tables()
        conn.execute("DELETE FROM priority_projects WHERE id = ?", (proj_id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/priority-projects/<int:proj_id>/complete', methods=['POST'])
def api_complete_priority_project(proj_id):
    """Move a project to the archive with today's date."""
    try:
        from datetime import date as _date_cls
        conn = _ensure_priority_tables()
        row = conn.execute(
            "SELECT * FROM priority_projects WHERE id = ?", (proj_id,)
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Project not found'}), 404
        d = dict(row)
        conn.execute("""
            INSERT INTO archived_projects
                (name, completed_date, priority, leader, process_type, next_gate, launch_date, objective, segment)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            d['name'], _date_cls.today().isoformat(), d['priority'],
            d['leader'], d['process_type'], d['next_gate'],
            d['launch_date'], d['objective'], d['segment']
        ))
        conn.execute("DELETE FROM priority_projects WHERE id = ?", (proj_id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/archived-projects', methods=['GET'])
def api_get_archived_projects():
    """Return all archived (completed) projects."""
    try:
        conn = _ensure_priority_tables()
        rows = conn.execute(
            "SELECT * FROM archived_projects ORDER BY completed_date DESC, id DESC"
        ).fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/archived-projects/<int:arch_id>/restore', methods=['POST'])
def api_restore_archived_project(arch_id):
    """Restore an archived project to the bottom of the active list."""
    try:
        conn = _ensure_priority_tables()
        row = conn.execute(
            "SELECT * FROM archived_projects WHERE id = ?", (arch_id,)
        ).fetchone()
        if not row:
            conn.close()
            return jsonify({'error': 'Archived project not found'}), 404
        d = dict(row)
        max_p = conn.execute(
            "SELECT COALESCE(MAX(priority), 0) FROM priority_projects"
        ).fetchone()[0]
        conn.execute("""
            INSERT INTO priority_projects
                (priority, name, leader, process_type, next_gate, launch_date, objective, segment)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            max_p + 1, d['name'], d['leader'], d['process_type'],
            d['next_gate'], d['launch_date'], d['objective'], d['segment']
        ))
        conn.execute("DELETE FROM archived_projects WHERE id = ?", (arch_id,))
        conn.commit()
        result = _build_priority_response(conn)
        conn.close()
        return jsonify({'success': True, 'projects': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# Gate Baseline and Change Log Endpoints

@app.route('/api/project/<project_name>/gate-baselines', methods=['GET'])
def api_get_gate_baselines(project_name):
    """Get all Gate baselines for a project"""
    try:
        baselines = db_manager.get_gate_baselines(project_name)
        return jsonify(baselines)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/project/<project_name>/gate-baselines', methods=['POST'])
def api_create_gate_baselines(project_name):
    """Create/update Gate baselines for a project"""
    try:
        data = request.json
        baselines = data.get('baselines', [])
        
        for baseline in baselines:
            db_manager.create_gate_baseline(
                project_name=project_name,
                gate_name=baseline['gate_name'],
                gate_id=baseline['gate_id'],
                baseline_date=baseline['baseline_date']
            )
        
        return jsonify({'success': True, 'count': len(baselines)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/project/<project_name>/gate-change-log', methods=['GET'])
def api_get_gate_change_log(project_name):
    """Get all Gate change log entries for a project"""
    try:
        changes = db_manager.get_gate_change_log(project_name)
        return jsonify(changes)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/project/<project_name>/gate-change-log', methods=['POST'])
def api_add_gate_change_log(project_name):
    """Add a Gate change log entry"""
    try:
        log_data = request.json
        log_data['project_name'] = project_name
        
        log_id = db_manager.add_gate_change_log(log_data)
        
        return jsonify({'success': True, 'log_id': log_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/project/<project_name>/gate-change-log/<gate_name>', methods=['DELETE'])
def api_delete_gate_change_log(project_name, gate_name):
    """Delete all change log entries for a specific Gate"""
    try:
        deleted_count = db_manager.delete_gate_change_log_by_gate(project_name, gate_name)
        
        return jsonify({'success': True, 'deleted_count': deleted_count})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/project/<project_name>/dependencies', methods=['GET'])
def api_get_dependencies(project_name):
    """Get all task dependencies for a project"""
    try:
        deps = db_manager.get_dependencies_by_project(project_name)
        return jsonify(deps)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/project/<project_name>/dependencies', methods=['POST'])
def api_create_dependency(project_name):
    """Create a task dependency"""
    try:
        data = request.json
        predecessor_id = data.get('predecessor_id')
        successor_id = data.get('successor_id')
        if not predecessor_id or not successor_id:
            return jsonify({'error': 'predecessor_id and successor_id required'}), 400
        dep_id = db_manager.create_dependency(project_name, predecessor_id, successor_id)
        return jsonify({'success': True, 'id': dep_id})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/disciplines')
def disciplines_page():
    return render_template('disciplines.html')


@app.route('/api/v1/admin/resource-teams', methods=['GET'])
def api_resource_teams_get():
    try:
        conn = db_manager.get_connection()
        rows = conn.execute(
            "SELECT id, team_name, owner_name, capacity_hrs_per_week FROM resource_teams ORDER BY owner_name"
        ).fetchall()
        conn.close()
        return jsonify([dict(r) for r in rows])
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/v1/admin/resource-teams/bulk', methods=['POST'])
def api_resource_teams_bulk():
    mappings = request.get_json(silent=True) or []
    if not isinstance(mappings, list):
        return jsonify({'error': 'Expected a JSON array of {owner_name, team_name}'}), 400
    try:
        conn = db_manager.get_connection()
        conn.execute("DELETE FROM resource_teams")
        count = 0
        for m in mappings:
            owner = (m.get('owner_name') or '').strip()
            team = (m.get('team_name') or '').strip()
            if owner and team:
                conn.execute(
                    """INSERT INTO resource_teams (owner_name, team_name, capacity_hrs_per_week)
                       VALUES (?, ?, ?)
                       ON CONFLICT(owner_name) DO UPDATE SET
                           team_name = excluded.team_name""",
                    (owner, team, m.get('capacity_hrs_per_week') or 37.5),
                )
                count += 1
        conn.commit()
        conn.close()
        return jsonify({'saved': count})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/discipline-resource-load')
def api_discipline_resource_load():
    import calendar
    from datetime import date as _date
    today = _date.today()

    # Build 13-month window: 1 past + current + 11 future
    months = []
    for i in range(-1, 12):
        raw_month = today.month - 1 + i
        y = today.year + raw_month // 12
        m = raw_month % 12 + 1
        months.append(_date(y, m, 1))

    # Load discipline mappings from dashboards.db (case-insensitive key)
    p_conn = db_manager.get_connection()
    mappings = p_conn.execute("SELECT TRIM(owner_name), team_name FROM resource_teams").fetchall()
    p_conn.close()
    owner_to_disc = {row[0].lower(): row[1] for row in mappings}

    # Count owners per discipline
    disc_owners = {}
    for owner_lower, disc in owner_to_disc.items():
        disc_owners[disc] = disc_owners.get(disc, 0) + 1

    # Load active tasks from dashboards.db
    db_path = config.DATABASE_PATH
    t_conn = sqlite3.connect(str(db_path))
    tasks = t_conn.execute("""
        SELECT TRIM(owner), start_date, end_date, status
        FROM tasks
        WHERE status IN ('In Process','Planned')
          AND start_date IS NOT NULL AND start_date != ''
          AND end_date   IS NOT NULL AND end_date   != ''
    """).fetchall()
    t_conn.close()

    # Compute monthly task counts per discipline
    disc_monthly = {}
    for owner_raw, start_str, end_str, status in tasks:
        disc = owner_to_disc.get((owner_raw or '').lower())
        if not disc:
            continue
        try:
            start = _date.fromisoformat(start_str[:10])
            end   = _date.fromisoformat(end_str[:10])
        except Exception:
            continue
        if disc not in disc_monthly:
            disc_monthly[disc] = {'In Process': [0]*len(months), 'Planned': [0]*len(months)}
        bucket = 'In Process' if status == 'In Process' else 'Planned'
        for i, ms in enumerate(months):
            last_day = calendar.monthrange(ms.year, ms.month)[1]
            me = _date(ms.year, ms.month, last_day)
            if start <= me and end >= ms:
                disc_monthly[disc][bucket][i] += 1

    # Inject PM overhead (+5/month per managed project) into discipline monthly counts
    PM_LOAD_PER_PROJECT = 5
    pm_conn = sqlite3.connect(str(db_path))
    pm_conn.row_factory = sqlite3.Row
    pm_proj_rows = pm_conn.execute("""
        SELECT TRIM(p.manager) AS manager,
               MIN(CASE WHEN t.status IN ('In Process','Planned') THEN t.start_date ELSE NULL END) AS proj_start,
               MAX(CASE WHEN t.status IN ('In Process','Planned') THEN t.end_date   ELSE NULL END) AS proj_end,
               SUM(CASE WHEN t.status IN ('In Process','Planned') THEN 1 ELSE 0 END) AS active_tasks
        FROM projects p
        LEFT JOIN tasks t ON t.project_id = p.id
        WHERE TRIM(COALESCE(p.manager,'')) != ''
        GROUP BY p.id, p.name, p.manager
        HAVING active_tasks > 0
    """).fetchall()
    pm_conn.close()

    for pr in pm_proj_rows:
        manager   = (pr['manager'] or '').strip()
        pstart_s  = pr['proj_start']
        pend_s    = pr['proj_end']
        if not manager or not pstart_s or not pend_s:
            continue
        # Resolve manager name → discipline (exact, then first-name)
        disc = owner_to_disc.get(manager.lower())
        if not disc:
            disc = owner_to_disc.get(manager.split()[0].lower())
        if not disc:
            continue
        try:
            pstart = _date.fromisoformat(pstart_s[:10])
            pend   = _date.fromisoformat(pend_s[:10])
        except Exception:
            continue
        if disc not in disc_monthly:
            disc_monthly[disc] = {'In Process': [0]*len(months), 'Planned': [0]*len(months)}
        for i, ms in enumerate(months):
            last_day = calendar.monthrange(ms.year, ms.month)[1]
            me = _date(ms.year, ms.month, last_day)
            if pstart <= me and pend >= ms:
                bucket = 'In Process' if ms <= today else 'Planned'
                disc_monthly[disc][bucket][i] += PM_LOAD_PER_PROJECT

    result = []
    for disc in sorted(disc_monthly):
        ip = disc_monthly[disc]['In Process']
        pl = disc_monthly[disc]['Planned']
        combined = [ip[i] + pl[i] for i in range(len(months))]
        peak = max(combined) if combined else 0
        peak_idx = combined.index(peak) if peak else 0
        owner_count = disc_owners.get(disc, 1)
        result.append({
            'name': disc,
            'owner_count': owner_count,
            'in_process': ip,
            'planned': pl,
            'combined': combined,
            'peak_count': peak,
            'peak_month': months[peak_idx].strftime('%Y-%m'),
        })

    result.sort(key=lambda x: x['peak_count'], reverse=True)
    return jsonify({
        'months': [m.strftime('%Y-%m') for m in months],
        'month_labels': [m.strftime("%b '%y") for m in months],
        'today_month': today.strftime('%Y-%m'),
        'disciplines': result,
    })


@app.route('/api/dependency/<int:dep_id>', methods=['DELETE'])
def api_delete_dependency(dep_id):
    """Delete a specific dependency"""
    try:
        success = db_manager.delete_dependency(dep_id)
        return jsonify({'success': success})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/project/<project_name>/dependencies/by-tasks', methods=['DELETE'])
def api_delete_dependency_by_tasks(project_name):
    """Delete a dependency between two specific tasks"""
    try:
        data = request.json
        success = db_manager.delete_dependency_by_tasks(data['predecessor_id'], data['successor_id'])
        return jsonify({'success': success})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/project/<project_name>/gate-sign-offs', methods=['GET'])
def api_get_gate_sign_offs(project_name):
    """Get all gate sign-offs for a project"""
    sign_offs = db_manager.get_gate_sign_offs(project_name)
    return jsonify(sign_offs)


@app.route('/api/project/<project_name>/gate-sign-offs', methods=['POST'])
def api_upsert_gate_sign_off(project_name):
    """Create or update a gate sign-off"""
    data = request.json
    required = ['gate_name', 'gate_id', 'sign_off_date', 'status']
    for field in required:
        if field not in data:
            return jsonify({'error': f'Missing required field: {field}'}), 400
    if data['status'] == 'Passed with Rework' and not data.get('rework_due_date'):
        return jsonify({'error': 'rework_due_date is required for Passed with Rework status'}), 400
    success = db_manager.upsert_gate_sign_off(
        project_name=project_name,
        gate_name=data['gate_name'],
        gate_id=data['gate_id'],
        sign_off_date=data['sign_off_date'],
        status=data['status'],
        rework_due_date=data.get('rework_due_date'),
        rework_sign_off_date=data.get('rework_sign_off_date')
    )
    return jsonify({'success': success})


@app.route('/api/project/<project_name>/gate-sign-off/<gate_name>', methods=['DELETE'])
def api_delete_gate_sign_off(project_name, gate_name):
    """Remove a gate sign-off"""
    success = db_manager.delete_gate_sign_off(project_name, gate_name)
    return jsonify({'success': success})


@app.route('/api/project/<project_name>/soft-delete', methods=['POST'])
def api_soft_delete_project(project_name):
    """Mark project as deleted (reversible within undo window)."""
    try:
        db_manager.soft_delete_project(project_name)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/project/<project_name>/restore', methods=['POST'])
def api_restore_project(project_name):
    """Restore a soft-deleted project."""
    try:
        db_manager.restore_project(project_name)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/project/<project_name>/rename', methods=['PUT'])
def api_rename_project(project_name):
    """Rename a project across all tables."""
    try:
        new_name = (request.json or {}).get('new_name', '').strip()
        if not new_name:
            return jsonify({'error': 'new_name is required'}), 400
        if new_name == project_name:
            return jsonify({'success': True})
        conn = db_manager.get_connection()
        if not conn.execute('SELECT id FROM projects WHERE name = ?', (project_name,)).fetchone():
            conn.close()
            return jsonify({'error': 'Project not found'}), 404
        if conn.execute('SELECT id FROM projects WHERE name = ?', (new_name,)).fetchone():
            conn.close()
            return jsonify({'error': f'A project named "{new_name}" already exists'}), 409
        conn.execute('UPDATE projects            SET name         = ? WHERE name         = ?', (new_name, project_name))
        conn.execute('UPDATE gate_sign_offs      SET project_name = ? WHERE project_name = ?', (new_name, project_name))
        conn.execute('UPDATE gate_baselines      SET project_name = ? WHERE project_name = ?', (new_name, project_name))
        conn.execute('UPDATE gate_change_log     SET project_name = ? WHERE project_name = ?', (new_name, project_name))
        conn.execute('UPDATE task_dependencies   SET project_name = ? WHERE project_name = ?', (new_name, project_name))
        conn.execute('UPDATE priority_projects   SET linked_dashboard = ? WHERE linked_dashboard = ?', (new_name, project_name))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'new_name': new_name})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/project/<project_name>', methods=['DELETE'])
def api_delete_project(project_name):
    """Delete a project and all its data"""
    try:
        success = db_manager.delete_project(project_name)
        if success:
            return jsonify({'success': True, 'message': f'Project "{project_name}" deleted'})
        else:
            return jsonify({'success': False, 'error': 'Project not found'}), 404
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


if __name__ == '__main__':
    print("=" * 80)
    print("Dashboard Generator Web Server")
    print("=" * 80)
    print(f"Server URL: http://localhost:{config.SERVER_PORT}")
    print(f"Database: {config.DATABASE_PATH}")
    print(f"Upload endpoint: /api/upload-excel")
    print("=" * 80)
    print("\nServer starts clean - upload Excel files via web interface")
    print("Press Ctrl+C to stop the server\n")
    
    app.run(
        host=config.SERVER_HOST,
        port=config.SERVER_PORT,
        debug=config.DEBUG_MODE
    )
