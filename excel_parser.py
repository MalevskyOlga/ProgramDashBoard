"""
Excel Parser for Dashboard Generator Web Server
Parses Excel files and extracts project and task data
"""

import openpyxl
from datetime import datetime, timedelta
import re
import config


class ExcelParser:
    def __init__(self):
        pass
    
    def parse_excel_file(self, file_path):
        """
        Parse Excel file and extract project info and tasks
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            tuple: (project_info dict, tasks list)
        """
        print(f"📖 Parsing Excel file: {file_path}")
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            # Extract project information
            project_info = {
                'name': str(ws['C3'].value or '').strip(),
                'manager': str(ws['C5'].value or '').strip(),
                'excel_filename': file_path
            }
            
            # If project name is empty, use filename
            if not project_info['name']:
                import os
                project_info['name'] = os.path.splitext(os.path.basename(file_path))[0]
            
            print(f"  ✓ Project: {project_info['name']}")
            print(f"  ✓ Manager: {project_info['manager']}")
            
            # Parse tasks starting from row 11
            tasks = []
            row_order = 1
            
            for row_num in range(config.EXCEL_DATA_START_ROW, ws.max_row + 1):
                row = ws[row_num]
                
                # Get values from columns
                ref_id = row[0].value  # Column A
                description = row[1].value  # Column B (Task Name)
                source = row[2].value  # Column C (Phase/Gate)
                assigned_to = row[3].value  # Column D (Owner)
                date_entered = row[4].value  # Column E (Start Date)
                status = row[5].value  # Column F (Status)
                due_date = row[6].value  # Column G (End Date)
                date_closed = row[7].value  # Column H
                result = row[8].value  # Column I
                
                # Skip empty rows
                if not description:
                    continue
                
                # Check if this is a Gate milestone (e.g., "Gate 4", "Gate 5")
                description_str = str(description or '').strip()
                is_gate_milestone = bool(re.match(r'^Gate\s*\d+\s*$', description_str, re.IGNORECASE))
                
                # Skip rows without owner (unless it's a gate milestone)
                if not assigned_to and not is_gate_milestone:
                    continue
                
                # Parse dates
                start_date_str = self._parse_date(date_entered)
                end_date_str = self._parse_date(due_date)
                
                # If end date is missing, set to 30 days from start
                if not end_date_str and start_date_str:
                    try:
                        start_dt = datetime.strptime(start_date_str, '%Y-%m-%d')
                        end_date_str = (start_dt + timedelta(days=30)).strftime('%Y-%m-%d')
                    except:
                        end_date_str = ''
                
                # Parse closed date
                closed_date_str = self._parse_date(date_closed) if date_closed else ''
                
                # Create task object
                task = {
                    'reference_id': str(ref_id or '').strip(),
                    'name': description_str,
                    'phase': str(source or '').strip(),
                    'owner': str(assigned_to or '').strip(),
                    'start_date': start_date_str,
                    'end_date': end_date_str,
                    'status': str(status or 'Planned').strip(),
                    'date_closed': closed_date_str,
                    'result': str(result or '').strip(),
                    'milestone': 1 if is_gate_milestone else 0,
                    'row_order': row_order
                }
                
                tasks.append(task)
                row_order += 1
            
            print(f"  ✓ Parsed {len(tasks)} tasks")
            
            wb.close()
            
            return project_info, tasks
            
        except Exception as e:
            print(f"  ✗ Error parsing Excel file: {e}")
            raise
    
    def _parse_date(self, date_value):
        """
        Parse date from Excel cell value
        
        Args:
            date_value: Date value from Excel cell
            
        Returns:
            str: Date in YYYY-MM-DD format, or empty string if invalid
        """
        if isinstance(date_value, datetime):
            return date_value.strftime('%Y-%m-%d')
        elif isinstance(date_value, str) and date_value:
            # Try to parse string date
            date_str = date_value.strip()
            
            # Try common date formats
            formats = ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d']
            
            for fmt in formats:
                try:
                    parsed_date = datetime.strptime(date_str, fmt)
                    return parsed_date.strftime('%Y-%m-%d')
                except ValueError:
                    continue
            
            # If no format matched, return empty
            return ''
        else:
            # Default to today for missing start dates
            return datetime.now().strftime('%Y-%m-%d')
